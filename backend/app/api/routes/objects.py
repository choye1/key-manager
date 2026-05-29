from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
from io import BytesIO


from io import BytesIO
from openpyxl import Workbook, load_workbook
from app.api.deps import get_current_user
from app.core.db import get_session
from app.crud import (
    create_object,
    get_issued_keys_for_object,
    get_object_by_code,
    get_objects,
)
from app.models import (
    IssuedKeyPublic,
    IssuedKeysResponse,
    Object,
    ObjectCreate,
    ObjectPublic,
    ObjectsPublic,
    ObjectType,
    User,
    UserRole,
)

router = APIRouter(prefix="/objects", tags=["objects"])


@router.post("", response_model=ObjectPublic)
def create_new_object(
    object_in: ObjectCreate,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    if current_user.role not in {UserRole.operator, UserRole.admin}:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions",
        )

    existing_object = get_object_by_code(session, object_in.code)
    if existing_object is not None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Object with this code already exists",
        )

    return create_object(session, object_in)


@router.get("", response_model=ObjectsPublic)
def read_objects(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    objects = get_objects(session)
    return ObjectsPublic(data=objects, count=len(objects))


@router.get("/export")
def export_objects(
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions",
        )
    objects = session.exec(select(Object)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Objects"

    ws.append(["code", "type", "is_active"])

    for obj in objects:
        ws.append([
            obj.code,
            obj.type.value,
            obj.is_active,
        ])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=objects.xlsx"
        },
    )

@router.post("/import")
def import_objects(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):

    if current_user.role != UserRole.admin:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions",
        )

    contents = file.file.read()
    wb = load_workbook(BytesIO(contents))
    ws = wb.active

    created = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        code, object_type, is_active = row

        existing = session.exec(
            select(Object).where(Object.code == code)
        ).first()

        if existing:
            continue

        obj = Object(
            code=code,
            type=ObjectType(object_type),
            is_active=bool(is_active),
        )

        session.add(obj)
        created += 1

    session.commit()

    return {
        "message": f"Imported {created} objects"
    }

@router.get("/{code}", response_model=ObjectPublic)
def read_object_by_code(
    code: str,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    obj = get_object_by_code(session, code)
    if obj is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Object not found",
        )

    return obj

@router.get("/{code}/issued-keys", response_model=IssuedKeysResponse)
def read_issued_keys_for_object(
    code: str,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    obj = get_object_by_code(session, code)
    if obj is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Object not found",
        )

    issued_keys = get_issued_keys_for_object(session, obj.id)

    return IssuedKeysResponse(
        object_code=obj.code,
        count=len(issued_keys),
        data=issued_keys,
    )

